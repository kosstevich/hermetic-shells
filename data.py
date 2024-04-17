from dataclasses import dataclass
from openpyxl import load_workbook
from datetime import datetime
import tabulate

@dataclass
class Cassete:   
    id: int             # Row number in the table        
    id1: int            # Measure number
    id2: int            # Id cassete
    name: str           # Cassete name
    coordinates: str    # Cassete coordinates
    date_time: datetime  
    companies: int      # Cassete age
    id_pen: int         # Number of experemental penal
    a_i131: float       # Activity I^131
    a_cs134: float 
    a_cs136: float 
    a_cs137: float 
    a_xe133: float 
    a_mn54: float
    status: int         # 0 - hermetic, 1 - non-hermetic, 2 - unknown, -1 - need to check it again

    def to_list(self)->list: return list([self.id, self.id1, self.id2, self.name, self.coordinates, 
                                self.date_time, self.companies, self.id_pen, self.a_i131, self.a_cs134, self.a_cs136, self.a_cs137, self.a_xe133, self.a_mn54, self.status])

class Data:
    def __init__(self, filename:str = "data/sample-data.xlsx", sheet_name:str = "Лист1"):
        self.input_data, self.cass_indexes = self._read(filename, sheet_name)
        self.clear_data = self._clear()

    def _read(self, filename:str, sheet_name:str):
        file = load_workbook(filename)
        sheet = file.get_sheet_by_name(sheet_name)
        rows = sheet.max_row
        input = []
        indexes = []

        for i in range(2, rows + 1):
            if sheet.cell(row=i, column=6).value == None:
                break
            cassete = Cassete(
                id = i,
                id1 = sheet.cell(row=i, column=1).value,
                id2 = sheet.cell(row=i, column=2).value,
                name = sheet.cell(row=i, column=3).value,
                coordinates = sheet.cell(row=i, column=5).value,
                date_time = sheet.cell(row=i, column=6).value,
                companies = sheet.cell(row=i, column=7).value,
                id_pen = sheet.cell(row=i, column=8).value,
                a_i131 = sheet.cell(row=i, column=9).value,
                a_cs134 = sheet.cell(row=i, column=10).value,
                a_cs136 = sheet.cell(row=i, column=13).value,
                a_xe133 = sheet.cell(row=i, column=14).value,
                a_cs137 = sheet.cell(row=i, column=11).value,
                a_mn54 = sheet.cell(row=i, column=12).value,
                status = 2
            )
            input.append(cassete)
            indexes.append(i)
        return input, indexes
    
    def _clear(self):
        clear_data = []
        for i in range(0,len(self.input_data)):
            if self.input_data[i].id1 != None and self.input_data[i].id2 != None:
                if self.input_data[i].a_i131 == "н/н":
                    self.input_data[i].a_i131 = 0
                if self.input_data[i].a_cs134 == "н/н":
                    self.input_data[i].a_cs134 = 0
                if self.input_data[i].a_cs136 == "н/н":
                    self.input_data[i].a_cs136 = 0
                if self.input_data[i].a_cs137 == "н/н":
                    self.input_data[i].a_cs137 = 0
                if self.input_data[i].a_xe133 == "н/н":
                    self.input_data[i].a_xe133 = 0
                if self.input_data[i].a_mn54 == "н/н":
                    self.input_data[i].a_mn54 = 0
                clear_data.append(self.input_data[i])

        return clear_data

    def _get_interval_by_range_id(self, left:int, right:int)->list:
        output = []
        # for i in range(left, right+1): output.append(self.input_data[self.cass_indexes.index(i)])
        i = 0 
        while self.clear_data[i].id1 <= right:
            if self.clear_data[i].id1 >= left:
                output.append(self.clear_data[i])
            i+=1
            if i==len(self.clear_data):
                break
        return output
    
    def divide_into_intervals(self, intervals:list)->list: 
        intervals_data = []
        for i in range(0,len(intervals)):
            intervals_data.append(self._get_interval_by_range_id(intervals[i][0],intervals[i][1]))
        return intervals_data

    def print(self, content:str):
        ls = []
        for i in range(0,len(self.__dict__[content])):
            ls.append(self.__dict__[content][i].to_list())
        print(tabulate.tabulate(ls))
